"""
Author: Perry Radau
Date: 2025-01-27
Brief description: CSV report generator for spending analysis
Dependencies: Python 3.6+
Usage: ReportGenerator class to create CSV reports from analysis results
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from analyzer import SpendingAnalyzer
from budget import BudgetManager
from transaction import Transaction


class ReportGenerator:
    """Generates CSV reports from spending analysis."""
    
    def __init__(self, analyzer: SpendingAnalyzer, output_dir: Path):
        """Initialize report generator.
        
        Args:
            analyzer: SpendingAnalyzer instance with transaction data
            output_dir: Directory to write CSV reports
        """
        self.analyzer = analyzer
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_monthly_summary(self, filename: str = "monthly_summary.csv",
                                 start_date: Optional[datetime] = None,
                                 end_date: Optional[datetime] = None) -> Path:
        """Generate monthly category summary report.
        
        Args:
            filename: Output filename
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated CSV file
        """
        monthly_data = self.analyzer.monthly_summary(start_date, end_date)
        
        # Collect all categories across all months
        all_categories = set()
        for month_data in monthly_data.values():
            all_categories.update(month_data.keys())
        all_categories = sorted(all_categories)
        
        output_path = self.output_dir / filename
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header row
            header = ['Month'] + all_categories + ['Total']
            writer.writerow(header)
            
            # Data rows
            for month in sorted(monthly_data.keys()):
                row = [month]
                month_total = 0.0
                
                for category in all_categories:
                    amount = monthly_data[month].get(category, 0.0)
                    row.append(f"{amount:.2f}")
                    month_total += amount
                
                row.append(f"{month_total:.2f}")
                writer.writerow(row)
        
        return output_path
    
    def generate_category_totals(self, filename: str = "category_totals.csv",
                                 start_date: Optional[datetime] = None,
                                 end_date: Optional[datetime] = None) -> Path:
        """Generate category totals report.
        
        Args:
            filename: Output filename
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated CSV file
        """
        totals = self.analyzer.category_totals(start_date, end_date)
        
        # Sort by amount descending
        sorted_totals = sorted(totals.items(), key=lambda x: x[1], reverse=True)
        
        output_path = self.output_dir / filename
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Category', 'Total Amount'])
            
            for category, amount in sorted_totals:
                writer.writerow([category, f"{amount:.2f}"])
        
        return output_path
    
    def generate_top_merchants(self, filename: str = "top_merchants.csv",
                              limit: int = 20,
                              start_date: Optional[datetime] = None,
                              end_date: Optional[datetime] = None) -> Path:
        """Generate top merchants report.
        
        Args:
            filename: Output filename
            limit: Maximum number of merchants to include
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated CSV file
        """
        top_merchants = self.analyzer.top_merchants(limit, start_date, end_date)
        
        output_path = self.output_dir / filename
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Merchant', 'Total Amount', 'Transaction Count'])
            
            for merchant, amount, count in top_merchants:
                writer.writerow([merchant, f"{amount:.2f}", count])
        
        return output_path
    
    def generate_spending_trends(self, filename: str = "spending_trends.csv") -> Path:
        """Generate spending trends report (month-over-month by category).
        
        Args:
            filename: Output filename
            
        Returns:
            Path to generated CSV file
        """
        trends = self.analyzer.spending_trends()
        
        # Collect all months across all categories
        all_months = set()
        for category_trends in trends.values():
            all_months.update(month for month, _ in category_trends)
        all_months = sorted(all_months)
        
        output_path = self.output_dir / filename
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header row
            header = ['Category'] + all_months
            writer.writerow(header)
            
            # Data rows
            for category in sorted(trends.keys()):
                row = [category]
                category_data = dict(trends[category])
                
                for month in all_months:
                    amount = category_data.get(month, 0.0)
                    row.append(f"{amount:.2f}")
                
                writer.writerow(row)
        
        return output_path
    
    def generate_transaction_list(self, filename: str = "transactions.csv",
                                 start_date: Optional[datetime] = None,
                                 end_date: Optional[datetime] = None) -> Path:
        """Generate detailed transaction list with categories.
        
        Args:
            filename: Output filename
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated CSV file
        """
        transactions = self.analyzer.filter_by_date_range(start_date, end_date)
        transactions = sorted(transactions, key=lambda t: t.date)
        
        output_path = self.output_dir / filename
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Date', 'Account Type', 'Description', 'Amount', 
                'Category', 'Subcategory', 'Credit Card'
            ])
            
            for transaction in transactions:
                writer.writerow([
                    transaction.date.strftime("%Y-%m-%d"),
                    transaction.account_type,
                    transaction.description,
                    f"{transaction.amount:.2f}",
                    transaction.category,
                    transaction.subcategory or "",
                    transaction.credit_card or ""
                ])
        
        return output_path
    
    def generate_budget_comparison(self, budget_manager: BudgetManager,
                                   filename: str = "budget_comparison.csv",
                                   start_date: Optional[datetime] = None,
                                   end_date: Optional[datetime] = None) -> Path:
        """Generate budget vs. actual spending comparison report.
        
        Args:
            budget_manager: BudgetManager instance with budget data
            filename: Output filename
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated CSV file
        """
        budget_comparison = self.analyzer.budget_vs_actual(
            budget_manager, start_date, end_date
        )
        
        # Sort by utilization (over budget first)
        sorted_comparison = sorted(
            budget_comparison.items(),
            key=lambda x: x[1]['utilization'],
            reverse=True
        )
        
        output_path = self.output_dir / filename
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Category', 'Budget', 'Actual', 'Difference', 
                'Utilization %', 'Status'
            ])
            
            for category, data in sorted_comparison:
                # Skip categories with no budget
                if data['budget'] == 0:
                    continue
                
                utilization = data['utilization']
                if utilization <= 80:
                    status = 'Under Budget'
                elif utilization <= 100:
                    status = 'On Track'
                elif utilization <= 120:
                    status = 'Over Budget'
                else:
                    status = 'Significantly Over'
                
                writer.writerow([
                    category,
                    f"{data['budget']:.2f}",
                    f"{data['actual']:.2f}",
                    f"{data['difference']:.2f}",
                    f"{utilization:.1f}",
                    status
                ])
        
        return output_path
    
    def export_to_json(self, filename: str = "transactions.json",
                      start_date: Optional[datetime] = None,
                      end_date: Optional[datetime] = None) -> Path:
        """Export transactions to JSON format.
        
        Args:
            filename: Output filename
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated JSON file
        """
        transactions = self.analyzer.filter_by_date_range(start_date, end_date)
        transactions = sorted(transactions, key=lambda t: t.date)
        
        # Convert transactions to dictionaries
        transaction_data = []
        for t in transactions:
            data = {
                'date': t.date.isoformat(),
                'account_type': t.account_type,
                'description': t.description,
                'amount': t.amount,
                'category': t.category,
                'subcategory': t.subcategory,
                'credit_card': t.credit_card,
                'tags': t.tags,
                'notes': t.notes
            }
            transaction_data.append(data)
        
        output_path = self.output_dir / filename
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump({
                'transactions': transaction_data,
                'count': len(transaction_data),
                'export_date': datetime.now().isoformat()
            }, f, indent=2, default=str)
        
        return output_path
    
    def export_to_pdf(self, filename: str = "summary.pdf",
                     start_date: Optional[datetime] = None,
                     end_date: Optional[datetime] = None) -> Optional[Path]:
        """Export summary report to PDF format.
        
        Note: Requires reportlab library. Returns None if not available.
        
        Args:
            filename: Output filename
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated PDF file, or None if reportlab not available
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return None
        
        output_path = self.output_dir / filename
        
        # Create PDF document
        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("Personal Finance Summary Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Summary statistics
        income_expenses = self.analyzer.income_vs_expenses(start_date, end_date)
        summary_data = [
            ['Metric', 'Amount'],
            ['Total Income', f"${income_expenses['income']:,.2f}"],
            ['Total Expenses', f"${income_expenses['expenses']:,.2f}"],
            ['Net', f"${income_expenses['net']:,.2f}"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 12))
        
        # Category totals
        category_totals = self.analyzer.category_totals(start_date, end_date)
        sorted_totals = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)
        
        category_data = [['Category', 'Total Amount']]
        for category, amount in sorted_totals:
            category_data.append([category, f"${amount:,.2f}"])
        
        category_table = Table(category_data)
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(Paragraph("Spending by Category", styles['Heading2']))
        story.append(category_table)
        
        # Build PDF
        doc.build(story)
        
        return output_path
    
    def export_to_excel(self, filename: str = "summary.xlsx",
                       start_date: Optional[datetime] = None,
                       end_date: Optional[datetime] = None) -> Optional[Path]:
        """Export reports to Excel format.
        
        Note: Requires openpyxl library. Returns None if not available.
        
        Args:
            filename: Output filename
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            
        Returns:
            Path to generated Excel file, or None if openpyxl not available
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return None
        
        output_path = self.output_dir / filename
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Personal Finance Summary Report"])
        ws_summary.append([])
        
        income_expenses = self.analyzer.income_vs_expenses(start_date, end_date)
        ws_summary.append(["Metric", "Amount"])
        ws_summary.append(["Total Income", income_expenses['income']])
        ws_summary.append(["Total Expenses", income_expenses['expenses']])
        ws_summary.append(["Net", income_expenses['net']])
        
        # Category totals sheet
        ws_categories = wb.create_sheet("Categories")
        ws_categories.append(["Category", "Total Amount"])
        
        category_totals = self.analyzer.category_totals(start_date, end_date)
        sorted_totals = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)
        for category, amount in sorted_totals:
            ws_categories.append([category, amount])
        
        # Transactions sheet
        ws_transactions = wb.create_sheet("Transactions")
        ws_transactions.append([
            "Date", "Account Type", "Description", "Amount", 
            "Category", "Subcategory", "Credit Card", "Tags", "Notes"
        ])
        
        transactions = self.analyzer.filter_by_date_range(start_date, end_date)
        transactions = sorted(transactions, key=lambda t: t.date)
        for t in transactions:
            ws_transactions.append([
                t.date.strftime("%Y-%m-%d"),
                t.account_type,
                t.description,
                t.amount,
                t.category,
                t.subcategory or "",
                t.credit_card or "",
                ", ".join(t.tags),
                t.notes or ""
            ])
        
        # Style headers
        for ws in [ws_summary, ws_categories, ws_transactions]:
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        wb.save(output_path)
        
        return output_path
    
    def generate_all_reports(self, start_date: Optional[datetime] = None,
                            end_date: Optional[datetime] = None,
                            budget_manager: Optional[BudgetManager] = None) -> Dict[str, Path]:
        """Generate all available reports.
        
        Args:
            start_date: Start date filter (inclusive)
            end_date: End date filter (inclusive)
            budget_manager: Optional BudgetManager for budget reports
            
        Returns:
            Dict mapping report names to file paths
        """
        reports = {}
        
        reports['monthly_summary'] = self.generate_monthly_summary(
            start_date=start_date, end_date=end_date
        )
        reports['category_totals'] = self.generate_category_totals(
            start_date=start_date, end_date=end_date
        )
        reports['top_merchants'] = self.generate_top_merchants(
            start_date=start_date, end_date=end_date
        )
        reports['spending_trends'] = self.generate_spending_trends()
        reports['transactions'] = self.generate_transaction_list(
            start_date=start_date, end_date=end_date
        )
        
        if budget_manager:
            reports['budget_comparison'] = self.generate_budget_comparison(
                budget_manager, start_date=start_date, end_date=end_date
            )
        
        return reports

